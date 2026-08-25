"""
add_extra_charts.py — build the CaptureVsBase sheet and the charts that read it, plus
the per-country variants, so `generate.py` reproduces the whole workbook.

WHY THIS EXISTS. Until 2026-08-25 the pipeline produced 23 sheets and 19 charts, while
the workbook actually in circulation had 24 and 62. The extra sheet (CaptureVsBase) and
the extra 43 charts had been added by hand in Excel and were in no script anywhere, so a
fresh generate.py run handed back a workbook missing two thirds of its exhibits and
nobody could tell without counting. This closes that gap: everything in the circulated
file is now built from the published CSVs.

WHAT IT ADDS
  CaptureVsBase          capture minus base, and capture as a percentage of base, per
                         country and technology, monthly, plus the rolling w1..w8 blocks
                         the monthly line charts read
  13 variant charts      per-country capture, capacity, intraday shape and cumulative
                         near-negative hours, cloned from the equivalents the pipeline
                         already builds and repointed at another country's columns
  37 monthly charts      one per country and technology in capture_vs_base.CHARTED

HOW IT RUNS. Late in the chain, AFTER every existing script, and purely additively: it
opens the finished workbook, appends parts and never rewrites one. That is deliberate.
The chain that produces the first 19 charts has been debugged over months and carries
several fixes whose reasons are recorded in those scripts; the cheapest way to keep all
of that true is not to touch it.

It never opens the workbook with openpyxl for writing, which would strip Power Query.
All edits are made to the raw package parts, the same way add_phase4_charts.py works.
"""
from __future__ import annotations

import os
import re
import shutil
import warnings
import zipfile

warnings.filterwarnings("ignore")
from lxml import etree
from openpyxl.utils import get_column_letter

import config as cfg
import capture_vs_base as cvb

ROOT = cfg.ROOT
WB = os.path.join(ROOT, "outputs", "HourlyPowerData.xlsx")
TEMPLATE = os.path.join(ROOT, "archive", "extra-charts-base-2026-08-25",
                        "HourlyPowerData_with-extra-charts.xlsx")

M = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG = "http://schemas.openxmlformats.org/package/2006/relationships"
CT = "http://schemas.openxmlformats.org/package/2006/content-types"
XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
CNS = "http://schemas.openxmlformats.org/drawingml/2006/chart"

SHEET_NAME = "CaptureVsBase"

# The Status sheet holds the rolling year labels in G2..N2 (w1..w8). The window blocks
# INDEX into a technology's percentage column at the offset those years imply, which is
# what makes every monthly chart roll in January without being rebuilt.
STATUS_W_COLS = [get_column_letter(7 + i) for i in range(8)]   # G..N

MONTH_ROWS = 12          # window blocks are 12 rows: one per calendar month
FIRST_DATA_ROW = 2


# ---------------------------------------------------------------------------
# CaptureVsBase worksheet
# ---------------------------------------------------------------------------
def _cell(ref, formula):
    return f'<c r="{ref}"><f>{formula}</f></c>'


def build_capture_vs_base_xml():
    """The whole sheet as worksheet XML: formulas only, no cached values.

    No cached values because there is nothing honest to cache — the figures depend on a
    Power Query refresh that has not happened yet. The workbook is set to recalculate on
    load (see set_full_calc_on_load), so Excel fills them the moment it opens, and a
    stale cache can never be shown as if it were current.
    """
    hdr, index = cvb.layout()
    n_cols = len(hdr)
    n_rows = len(cvb.MONTHS) + 1

    rows = []
    # header
    cells = "".join(
        f'<c r="{get_column_letter(i + 1)}1" t="inlineStr"><is><t xml:space="preserve">'
        f'{_esc(h)}</t></is></c>' for i, h in enumerate(hdr))
    rows.append(f'<row r="1" spans="1:{n_cols}">{cells}</row>')

    pairs = cvb.pair_columns()
    windows = cvb.window_columns()

    for r in range(FIRST_DATA_ROW, n_rows + 1):
        cells = [_cell(f"A{r}", f"{cvb.CAPTURE_SHEET}!A{r}")]
        for country, tech, diff_h, pct_h in pairs:
            src = cvb.source_sheet(country)
            sc = get_column_letter(cvb.source_col_index(country, tech))
            pc = get_column_letter(cvb.price_col_index(country))
            guard = f'OR({src}!{sc}{r}="",{cvb.PRICE_SHEET}!${pc}{r}="")'
            # A missing month is NA(), not zero and not blank: NA() is the one value a
            # line chart leaves as a gap rather than drawing through, so a market that
            # did not report a technology in a month shows nothing instead of a line
            # dropping to the axis and back.
            rows_diff = (f'IF({guard},NA(),ROUND({src}!{sc}{r}-'
                         f'{cvb.PRICE_SHEET}!${pc}{r},2))')
            rows_pct = (f'IF({guard},NA(),ROUND({src}!{sc}{r}/'
                        f'{cvb.PRICE_SHEET}!${pc}{r}*100,1))')
            cells.append(_cell(f"{get_column_letter(index[diff_h])}{r}", rows_diff))
            cells.append(_cell(f"{get_column_letter(index[pct_h])}{r}", rows_pct))

        if r <= FIRST_DATA_ROW + MONTH_ROWS - 1:
            mcol = get_column_letter(index[cvb.MONTH_LABEL_HEADER])
            label = cvb.MONTH_LABELS[r - FIRST_DATA_ROW]
            cells.append(f'<c r="{mcol}{r}" t="inlineStr"><is><t>{label}</t></is></c>')
            for country, tech, wh in windows:
                pct_col = get_column_letter(index[f"{country}_{tech} {cvb.PCT}"])
                for i, h in enumerate(wh):
                    wcol = get_column_letter(index[h])
                    scol = STATUS_W_COLS[i]
                    f = (f'INDEX(${pct_col}${FIRST_DATA_ROW}:${pct_col}${n_rows},'
                         f'(Status!{scol}$2-{cfg.START_YEAR})*12+ROW()-1)')
                    cells.append(_cell(f"{wcol}{r}", f))
        rows.append(f'<row r="{r}" spans="1:{n_cols}">{"".join(cells)}</row>')

    dim = f"A1:{get_column_letter(n_cols)}{n_rows}"
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<worksheet xmlns="{M}" xmlns:r="{R}">'
        f'<dimension ref="{dim}"/>'
        '<sheetViews><sheetView workbookViewId="0"/></sheetViews>'
        '<sheetFormatPr defaultRowHeight="15"/>'
        f'<sheetData>{"".join(rows)}</sheetData>'
        '<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" '
        'header="0.3" footer="0.3"/>'
        '</worksheet>'
    ).encode("utf-8")


def _esc(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


# ---------------------------------------------------------------------------
# package helpers
# ---------------------------------------------------------------------------
def read_parts(path):
    z = zipfile.ZipFile(path)
    parts = {i.filename: z.read(i.filename) for i in z.infolist()}
    order = [i.filename for i in z.infolist()]
    z.close()
    return parts, order


def write_parts(path, parts, order):
    tmp = path + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as z:
        for name in order:
            z.writestr(name, parts[name])
    shutil.move(tmp, path)


def sheet_part_for(parts, name):
    """Worksheet part path for a sheet name, via workbook.xml and its rels."""
    wb = etree.fromstring(parts["xl/workbook.xml"])
    rels = etree.fromstring(parts["xl/_rels/workbook.xml.rels"])
    rmap = {r.get("Id"): r.get("Target") for r in rels}
    for s in wb.findall(f".//{{{M}}}sheet"):
        if s.get("name") == name:
            t = rmap[s.get(f"{{{R}}}id")]
            return "xl/" + t.lstrip("/")
    return None


def max_sheet_id(parts):
    wb = etree.fromstring(parts["xl/workbook.xml"])
    return max(int(s.get("sheetId")) for s in wb.findall(f".//{{{M}}}sheet"))


def next_rel_id(rels_xml):
    ids = [int(m) for m in re.findall(rb'Id="rId(\d+)"', rels_xml)]
    return max(ids) + 1 if ids else 1


def add_content_type(parts, partname, content_type):
    ct = parts["[Content_Types].xml"].decode()
    if f'PartName="/{partname}"' in ct:
        return
    ct = ct.replace("</Types>",
                    f'<Override PartName="/{partname}" ContentType="{content_type}"/></Types>')
    parts["[Content_Types].xml"] = ct.encode()


def set_full_calc_on_load(parts):
    """Force a recalculation when the workbook opens.

    Every cell this script writes is a formula with no cached value. Without this flag
    Excel shows an empty grid until something triggers a calculation, and the charts
    that read those cells come up blank — which looks exactly like a data failure.
    """
    xml = parts["xl/workbook.xml"].decode()
    if "<calcPr" in xml:
        xml = re.sub(r"<calcPr[^/>]*/>", '<calcPr calcId="191029" fullCalcOnLoad="1"/>', xml)
    else:
        xml = xml.replace("</workbook>",
                          '<calcPr calcId="191029" fullCalcOnLoad="1"/></workbook>')
    parts["xl/workbook.xml"] = xml.encode()


def append_sheet(parts, order, name, sheet_xml):
    """Append a worksheet as the LAST sheet in document order.

    Last, because every ExternalData_1 defined name is scoped by a localSheetId that is
    the 0-based document-order index of its sheet. Inserting anywhere else renumbers
    them all and silently repoints each Power Query at the wrong tab.
    """
    existing = [n for n in parts if re.match(r"xl/worksheets/sheet\d+\.xml$", n)]
    nums = [int(re.search(r"(\d+)", os.path.basename(n)).group(1)) for n in existing]
    num = max(nums) + 1
    part = f"xl/worksheets/sheet{num}.xml"
    parts[part] = sheet_xml
    order.append(part)

    rels = parts["xl/_rels/workbook.xml.rels"]
    rid = f"rId{next_rel_id(rels)}"
    rels = rels.replace(
        b"</Relationships>",
        (f'<Relationship Id="{rid}" Type="{R}/worksheet" '
         f'Target="worksheets/sheet{num}.xml"/></Relationships>').encode())
    parts["xl/_rels/workbook.xml.rels"] = rels

    wb = parts["xl/workbook.xml"].decode()
    sid = max_sheet_id(parts) + 1
    wb = wb.replace("</sheets>",
                    f'<sheet name="{name}" sheetId="{sid}" r:id="{rid}"/></sheets>')
    parts["xl/workbook.xml"] = wb.encode()

    add_content_type(parts, part,
                     "application/vnd.openxmlformats-officedocument."
                     "spreadsheetml.worksheet+xml")
    return part


# ---------------------------------------------------------------------------
# charts
# ---------------------------------------------------------------------------
# The 13 per-country variants, by template chart part, with the caption the Charts tab
# carries above each. Their references point at Fig5_Window, Fig9_Window and Line_Window,
# whose layouts this change does not touch, so they are copied as they are.
VARIANTS = [
    (14, "France — capture price vs baseload by technology"),
    (15, "Spain — capture price vs baseload by technology"),
    (16, "Italy — capture price vs baseload by technology"),
    (17, "France — intraday price shape, indexed to daily mean"),
    (18, "Fig 9 — Installed generation capacity by technology (Spain)"),
    (19, "Fig 9 — Installed generation capacity by technology (Portugal)"),
    (20, "Fig 9 — Installed generation capacity by technology (France)"),
    (21, "Fig 9 — Installed generation capacity by technology (Italy)"),
    (22, "Portugal — intraday price shape, indexed to daily mean"),
    (23, "Italy — intraday price shape, indexed to daily mean"),
    (54, "Portugal — cumulative near-negative-price hours through the year"),
    (55, "France — cumulative near-negative-price hours through the year"),
    (56, "Italy — cumulative near-negative-price hours through the year"),
]

MONTHLY_TEMPLATE = 24        # the chart every monthly capture exhibit is cloned from

# Series names come from the Status sheet's rolling year cells, which is the convention
# every chart in this workbook already uses.
#
# ONE THING NOT COPIED FROM THE CIRCULATED FILE. Its hand-built charts took the eighth
# label from Status!$P$2, a helper cell holding =N2&" YTD". In the workbook this pipeline
# builds, P2 is `health_note` — part of the status table — so copying that reference
# verbatim would print a health message in the legend where the year should be. These
# charts point at Status!$N$2 like every other chart here. The cost is that the current
# year reads "2026" rather than "2026 YTD" after a refresh, which is pre-existing
# behaviour across all nineteen original charts, not something introduced here.
SERIES_NAME_CELLS = [f"Status!${get_column_letter(7 + i)}$2" for i in range(8)]

CAPTION_COLS = [0, 11]       # the Charts tab is a two-column grid: A and L
ROW_STEP = 20                # rows between one chart's caption and the next


def _strip_caches(xml):
    """Remove cached chart points, keeping the cache elements' shape.

    A cached point is a number Excel draws before it has recalculated. Carrying one over
    from a chart that now reads a DIFFERENT column means the exhibit renders real-looking
    figures that belong to another technology until the first refresh lands. Better to
    draw nothing for a moment than the wrong thing convincingly.
    """
    xml = re.sub(r"<c:pt idx=\"\d+\"><c:v>[^<]*</c:v></c:pt>", "", xml)
    # AND the count with them. A cache that declares ptCount="52" while holding no
    # points is not an empty cache, it is an inconsistent one, and Excel offers to
    # Recover the file rather than opening it. Caught by opc_validate.py on the first
    # build that used this function, which is exactly what that check is for.
    xml = re.sub(r'<c:ptCount val="\d+"/>', '<c:ptCount val="0"/>', xml)
    return xml


def _strip_user_shapes(xml):
    """Remove the <c:userShapes> reference from a cloned chart.

    THE FAULT THIS FIXES, found by opening the file in Excel on 2026-08-25. Every chart
    in the source workbook carries <c:userShapes r:id="rId1"/>, pointing through its own
    .rels file at a chart-drawing part. Cloning the chart XML without also copying that
    rels file leaves the reference dangling, and Excel answers a dangling relationship
    with "We found a problem with some content ... do you want us to try to recover",
    which is the one outcome this project treats as unacceptable because recovering
    strips Power Query.

    Every structural check passed on that file: the XML was well-formed, every part was
    declared in [Content_Types].xml, the package joins were consistent and the chart
    ranges were in bounds. None of them looked INSIDE a chart part for a relationship id
    and asked whether it resolved. opc_validate.py now does.

    Stripping rather than copying, because these overlays are empty — no shapes, no text
    runs, just a zero-size anchor left behind by Excel's editor. Nothing visible is lost.
    """
    return re.sub(r"<c:userShapes[^>]*/>", "", xml)


def _repoint(xml, mapping):
    """Rewrite every <c:f> reference through `mapping` (old -> new)."""
    def repl(m):
        ref = m.group(1)
        return f"<c:f>{mapping.get(ref, ref)}</c:f>"
    return re.sub(r"<c:f>([^<]+)</c:f>", repl, xml)


def build_monthly_chart(template_xml, country, tech, index):
    """One monthly capture chart, pointed at this country and technology's window block."""
    xml = template_xml
    month_col = get_column_letter(index[cvb.MONTH_LABEL_HEADER])
    mapping = {}
    # categories: the month-label column
    old_cat = re.findall(r"<c:cat><c:strRef><c:f>([^<]+)</c:f>", xml)
    for oc in set(old_cat):
        mapping[oc] = f"{SHEET_NAME}!${month_col}${FIRST_DATA_ROW}:${month_col}$" \
                      f"{FIRST_DATA_ROW + MONTH_ROWS - 1}"
    # values: the eight rolling window columns
    old_vals = re.findall(r"<c:val><c:numRef><c:f>([^<]+)</c:f>", xml)
    for i, ov in enumerate(old_vals):
        col = get_column_letter(index[f"{country}_{tech} w{i + 1}"])
        mapping[ov] = f"{SHEET_NAME}!${col}${FIRST_DATA_ROW}:${col}$" \
                      f"{FIRST_DATA_ROW + MONTH_ROWS - 1}"
    # series names
    old_tx = re.findall(r"<c:tx><c:strRef><c:f>([^<]+)</c:f>", xml)
    for i, ot in enumerate(old_tx):
        if i < len(SERIES_NAME_CELLS):
            mapping[ot] = SERIES_NAME_CELLS[i]
    return _strip_user_shapes(_strip_caches(_repoint(xml, mapping)))


# ---------------------------------------------------------------------------
# a country variant of an existing chart
# ---------------------------------------------------------------------------
# Great Britain needs the same four exhibits every other market has. Each is built by
# taking the chart another country already has and moving its column references across,
# which is how every country variant in this workbook has been made since Phase 4.
#
# (template chart, the CSV behind the tab it reads, the country it currently shows,
#  the caption for the new one)
UK_VARIANTS = [
    (15, "fig5_capture_window", "ES",
     "United Kingdom — capture price vs baseload by technology"),
    (18, "fig9_capacity_window", "ES",
     "Fig 9 — Installed generation capacity by technology (United Kingdom)"),
    (17, "line_windows", "FR",
     "United Kingdom — intraday price shape, indexed to daily mean"),
    (55, "line_windows", "FR",
     "United Kingdom — cumulative near-negative-price hours through the year"),
]


def _csv_header(stem):
    import csv
    for base in (os.path.join(ROOT, "published", "charts"),
                 os.path.join(cfg.OUTPUT_DIR, "csv", "charts")):
        p = os.path.join(base, f"{stem}.csv")
        if os.path.exists(p):
            with open(p, newline="", encoding="utf-8") as f:
                return next(csv.reader(f))
    return []


def _country_letter_map(stem, src, dst):
    """Column-letter substitutions that move a chart from one country to another.

    Built by NAME, not by arithmetic: each source column's header has its country token
    swapped and the result is looked up in the same header row. Offsetting by a fixed
    number of columns would be shorter and would break the first time a table's block
    width changed, silently pointing a chart at a neighbour.
    """
    header = _csv_header(stem)
    if not header:
        return {}
    pos = {h: i + 1 for i, h in enumerate(header)}
    out = {}
    for h, i in pos.items():
        # "ES_w3" -> "GB_w3"; "i_FR_w3" -> "i_GB_w3"
        new = re.sub(rf"(^|_){src}(_)", rf"\g<1>{dst}\g<2>", h)
        if new != h and new in pos:
            out[get_column_letter(i)] = get_column_letter(pos[new])
    return out


def build_country_variant(template_xml, stem, src, dst):
    """Repoint every reference in a chart from one country's columns to another's.

    THIS FUNCTION FAILED SILENTLY AND SHIPPED FOUR WRONG CHARTS (found by review,
    2026-08-25). The substitution pattern was written `r'\\$([A-Z]{1,3})\\$'`, which in a
    raw string is a literal backslash followed by a dollar — a sequence that never occurs
    in an Excel reference. So every reference passed through untouched, and four charts
    captioned "United Kingdom" were emitted plotting Spain's and France's columns. The
    letter map was correct; only the rewrite was dead, which is why checking the map
    looked like checking the function.

    Nothing caught it. The package was valid, the chart count was right, the ranges were
    in bounds, no column had moved, and the deck matched its spec. Wrong data under a
    right title is invisible to every positional check, which is why this now VERIFIES
    ITS OWN WORK and raises rather than returning a chart it failed to repoint.
    """
    letters = _country_letter_map(stem, src, dst)
    if not letters:
        raise SystemExit(
            f"build_country_variant: no {src} -> {dst} column mapping for {stem}. "
            f"Refusing to emit a chart that would carry {src}'s data under a {dst} label.")

    def repl(m):
        ref = m.group(1)
        # Substituting inside the reference text handles multi-area references
        # ("(Sheet!$W$2:$W$7,Sheet!$W$9)") without having to parse them apart.
        def one(mm):
            return f"${letters.get(mm.group(1), mm.group(1))}$"
        return f"<c:f>{re.sub(r'\$([A-Z]{1,3})\$', one, ref)}</c:f>"

    xml = re.sub(r"<c:f>([^<]+)</c:f>", repl, template_xml)

    # A repointed chart MUST differ from the one it was cloned from, and must no longer
    # mention any source-country column. Both halves matter: the first catches a dead
    # substitution, the second catches a partial one that moved some references and left
    # others pointing at the country whose chart this used to be.
    if xml == template_xml:
        raise SystemExit(
            f"build_country_variant: repointing {src} -> {dst} on {stem} changed nothing. "
            f"The chart would have shipped as {src}'s data under a {dst} label.")
    leftover = sorted({c for c in letters
                       if re.search(rf"\${c}\$\d", xml)})
    if leftover:
        raise SystemExit(
            f"build_country_variant: {src} -> {dst} on {stem} left column(s) "
            f"{', '.join(leftover)} still pointing at {src}.")
    mapping = {}
    for i, ot in enumerate(re.findall(r"<c:tx><c:strRef><c:f>([^<]+)</c:f>", xml)):
        if i < len(SERIES_NAME_CELLS):
            mapping[ot] = SERIES_NAME_CELLS[i]
    return _strip_user_shapes(_strip_caches(_repoint(xml, mapping)))


# ---------------------------------------------------------------------------
# hydro reservoir tracker
# ---------------------------------------------------------------------------
HYDRO_TAB = "HydroWindow"
HYDRO_CSV = "hydro_window"
EXTRA_CAPTURE_TAB = "CaptureMonthlyExtra"
EXTRA_CAPTURE_CSV = "capture_monthly_extra"

HYDRO_TEMPLATE = os.path.join(ROOT, "archive", "extra-charts-base-2026-08-25",
                              "Hydro_Tracker_reference.xlsx")
HYDRO_TEMPLATE_CHART = "xl/charts/chart10.xml"     # the Norway band-and-lines combo

HYDRO_WEEKS = 53
HYDRO_FIRST_ROW = 2

# The band chart draws five year-lines over the historic range, which is what the
# tracker's own charts show: the last four complete years plus the current one. Those
# are window slots w4..w8, so the slots roll every January and the chart does not.
HYDRO_LINE_SLOTS = [4, 5, 6, 7, 8]


def _hydro_labels():
    """[(column label, caption)] for every hydro chart, in build order."""
    out = []
    for key, _area, name in cfg.HYDRO_RESERVOIR_ZONES:
        out.append((key, f"{name} — hydro reservoir fill by week, against the "
                         f"historic range (TWh)"))
    for country in cfg.PUMPED_ONLY:
        name = cfg.COUNTRIES[country]["name"]
        # Named as pumped storage, never as reservoir. It is a weekly FLOW from a pumped
        # fleet, not a stored volume, and these two markets publish no reservoir series
        # at all — so a caption that blurred them would invite the chart to be read as
        # the thing it is standing in for.
        out.append((f"{country}pump",
                    f"{name} — pumped-storage generation by week, against the "
                    f"historic range (GWh)"))
    return out


BASE_CHART_COUNT = 19        # what the original chain builds, before anything here


def expected_chart_total():
    """How many chart parts the finished workbook should hold.

    Asked by check_consistency rather than recomputed there, so the assertion cannot
    drift from the build. The hydro count is data-dependent — a zone that published
    nothing gets no chart — so it is counted the same way the builder counts it.
    """
    hcols = _hydro_columns()
    hydro = sum(1 for label, _cap in _hydro_labels() if f"{label}_min" in hcols)
    return (BASE_CHART_COUNT + len(VARIANTS) + len(UK_VARIANTS)
            + sum(len(v) for v in cvb.CHARTED.values()) + hydro)


def build_hydro_chart(template_xml, label, hcols):
    """One band-and-lines chart, repointed from the tracker's Norway original."""
    xml = template_xml
    mapping = {}
    old_vals = re.findall(r"<c:val><c:numRef><c:f>([^<]+)</c:f>", xml)
    old_tx = re.findall(r"<c:tx><c:strRef><c:f>([^<]+)</c:f>", xml)

    # The template's series order is: area min, area range, then five year lines, then
    # the long-run average. Keeping that order is what preserves the stacked band, since
    # the invisible minimum has to be the first area series for the range to sit on it.
    wanted = [f"{label}_min", f"{label}_range"]
    wanted += [cfg.wcol(label, i) for i in HYDRO_LINE_SLOTS]
    wanted += [f"{label}_avg"]

    last = HYDRO_FIRST_ROW + HYDRO_WEEKS - 1
    for ov, want in zip(old_vals, wanted):
        col = get_column_letter(hcols[want])
        mapping[ov] = f"{HYDRO_TAB}!${col}${HYDRO_FIRST_ROW}:${col}${last}"
    for ot, want in zip(old_tx, wanted):
        col = get_column_letter(hcols[want])
        mapping[ot] = f"{HYDRO_TAB}!${col}$1"
    # The five year lines read their legend text from the rolling Status cells, so a new
    # January relabels them without the chart being rebuilt. Min, range and average keep
    # their header-cell names, which never change.
    for i, slot in enumerate(HYDRO_LINE_SLOTS):
        ot = old_tx[2 + i]
        mapping[ot] = f"Status!${get_column_letter(6 + slot)}$2"
    return _strip_user_shapes(_strip_caches(_repoint(xml, mapping)))


def _hydro_columns():
    """1-based column numbers on the HydroWindow tab, from the published CSV header."""
    import csv
    path = os.path.join(ROOT, "published", "charts", f"{HYDRO_CSV}.csv")
    if not os.path.exists(path):
        path = os.path.join(cfg.OUTPUT_DIR, "csv", "charts", f"{HYDRO_CSV}.csv")
    if not os.path.exists(path):
        return {}
    with open(path, newline="", encoding="utf-8") as f:
        header = next(csv.reader(f))
    return {h: i + 1 for i, h in enumerate(header)}


def empty_sheet_xml():
    """A worksheet with nothing in it, ready for Power Query to load into.

    Loading a query into EMPTY cells fills them in place with no shift, so a chart
    pointed at those cells keeps reading them. Seeding the target first is what makes
    Power Query push the existing content aside and strand the chart on a stale copy.
    """
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<worksheet xmlns="{M}" xmlns:r="{R}">'
        '<dimension ref="A1"/>'
        '<sheetViews><sheetView workbookViewId="0"/></sheetViews>'
        '<sheetFormatPr defaultRowHeight="15"/><sheetData/>'
        '<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" '
        'header="0.3" footer="0.3"/>'
        '</worksheet>').encode("utf-8")


# The five CROSS-MARKET charts: one series per country, so a sixth market means a sixth
# series rather than a new chart. They were missed when GB was added everywhere else,
# because "add a country" had meant "add a column block" up to that point.
#
# (chart part, sheet, GB's column, first/last data row, the caption's country name)
CROSS_MARKET_SERIES = [
    ("chart1.xml", "Line_Window", "f1_GB_w", 2, 8),
    ("chart3.xml", "Line_Window", "f3_GB_w", 2, 8),
    ("chart14.xml", "G1_SolarPeak", "GB_qavg", 2, 6210),
    ("chart16.xml", "A_MonthPrice", "GB", 2, 205),
    ("chart17.xml", "B_Penetration", "GB", 2, 205),
]

# Country colours in this workbook are 1F3864, 8A1E41, CC9F53, 2E7D8A and 3D664A. TEAL
# from the Redburn palette is unused and distinct from all five.
SIXTH_COUNTRY_COLOUR = "5FA1AD"


def _sheet_column(sheet, header_name):
    """1-based column of a named header on a query-loaded tab."""
    stem = {"Line_Window": "line_windows", "G1_SolarPeak": "g1_solar_peakhour",
            "A_MonthPrice": "figA_monthly_price",
            "B_Penetration": "figB_penetration"}[sheet]
    import csv as _csv
    for base in (os.path.join(cfg.OUTPUT_DIR, "csv", "charts"),
                 os.path.join(ROOT, "published", "charts")):
        f = os.path.join(base, f"{stem}.csv")
        if os.path.exists(f):
            with open(f, newline="", encoding="utf-8") as fh:
                hdr = next(_csv.reader(fh))
            if header_name in hdr:
                return hdr.index(header_name) + 1
    return None


def append_country_series(parts, part_name, sheet, header_name, r0, r1, label):
    """Add one more country series to a cross-market chart, by cloning the last one.

    Cloning rather than composing, so the new series inherits the chart's own idiom —
    marker settings, smoothing, axis ids, the lot — and only the things that must differ
    are changed: which column it reads, what it is called, and its colour.
    """
    path = f"xl/charts/{part_name}"
    if path not in parts:
        return f"{part_name}: not in the workbook"
    xml = parts[path].decode()
    col = _sheet_column(sheet, header_name)
    if col is None:
        return f"{part_name}: {sheet} has no column named {header_name!r}"
    letter = get_column_letter(col)
    ref = f"{sheet}!${letter}${r0}:${letter}${r1}"
    if ref in xml:
        return None                                   # already added; idempotent re-run

    sers = re.findall(r"<c:ser>.*?</c:ser>", xml, re.S)
    if not sers:
        return f"{part_name}: no series to clone"
    clone = sers[-1]
    n = len(sers)

    new = re.sub(r'<c:idx val="\d+"/>', f'<c:idx val="{n}"/>', clone, count=1)
    new = re.sub(r'<c:order val="\d+"/>', f'<c:order val="{n}"/>', new, count=1)
    new = re.sub(r"(<c:val><c:numRef><c:f>)[^<]+(</c:f>)", rf"\g<1>{ref}\g<2>", new, count=1)
    # The series NAME is a literal, matching every other series on these charts. That is
    # the project's own idiom (CHARTS.md): a name read from a cell can be reverted by a
    # Power Query refresh, a literal cannot.
    new = re.sub(r"<c:tx>.*?</c:tx>", f"<c:tx><c:v>{_esc(label)}</c:v></c:tx>",
                 new, count=1, flags=re.S)
    new = re.sub(r'<a:srgbClr val="[0-9A-Fa-f]{6}"/>',
                 f'<a:srgbClr val="{SIXTH_COUNTRY_COLOUR}"/>', new)
    new = _strip_caches(new)
    # A cloned series keeps the uniqueId of the one it came from, and two series sharing
    # one is a duplicate identity Excel objects to.
    new = re.sub(r'<c:extLst>.*?</c:extLst>', "", new, flags=re.S)

    xml = xml.replace(clone, clone + new, 1)
    parts[path] = xml.encode()
    return None


def _next_chart_num(parts):
    nums = [int(re.search(r"chart(\d+)\.xml$", n).group(1))
            for n in parts if re.match(r"xl/charts/chart\d+\.xml$", n)]
    return max(nums) + 1 if nums else 1


def _next_drawing_num(parts):
    nums = [int(re.search(r"drawing(\d+)\.xml$", n).group(1))
            for n in parts if re.match(r"xl/drawings/drawing\d+\.xml$", n)]
    return max(nums) + 1 if nums else 1


def _anchor(col, row, rid, shape_id, name):
    """A oneCellAnchor for a chart on the Charts tab, sized like its neighbours."""
    return (
        f'<xdr:oneCellAnchor xmlns:xdr="{XDR}" '
        f'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
        f'xmlns:c="{CNS}" xmlns:r="{R}">'
        f'<xdr:from><xdr:col>{col}</xdr:col><xdr:colOff>0</xdr:colOff>'
        f'<xdr:row>{row}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>'
        f'<xdr:ext cx="6858000" cy="3429000"/>'
        f'<xdr:graphicFrame macro=""><xdr:nvGraphicFramePr>'
        f'<xdr:cNvPr id="{shape_id}" name="{name}"/><xdr:cNvGraphicFramePr/>'
        f'</xdr:nvGraphicFramePr>'
        f'<xdr:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></xdr:xfrm>'
        f'<a:graphic><a:graphicData '
        f'uri="http://schemas.openxmlformats.org/drawingml/2006/chart">'
        f'<c:chart r:id="{rid}"/></a:graphicData></a:graphic>'
        f'</xdr:graphicFrame><xdr:clientData/></xdr:oneCellAnchor>')


def _caption_row(row, col_letter, text):
    """A caption cell as its own <row>, navy bold, matching the existing captions."""
    return (f'<row r="{row}"><c r="{col_letter}{row}" t="inlineStr" s="0">'
            f'<is><t xml:space="preserve">{_esc(text)}</t></is></c></row>')


def add_charts(parts, order, tparts, charts_part, index):
    """Clone the variant charts and build the monthly ones, then anchor and caption them.

    Everything is APPENDED: new chart parts take numbers above the highest in use, new
    anchors go on the end of the Charts tab's existing drawing, and new captions occupy
    rows below the last one. Nothing already in the workbook is renumbered or moved, so
    every reference that worked before this script ran still works after it.
    """
    drawing_rels_part = "xl/drawings/_rels/drawing10.xml.rels"
    drawing_part = "xl/drawings/drawing10.xml"
    # The Charts tab's drawing, found through its own rels rather than assumed.
    srels = parts[f"xl/worksheets/_rels/{os.path.basename(charts_part)}.rels"].decode()
    m = re.search(r'Target="\.\./(drawings/drawing\d+\.xml)"', srels)
    if m:
        drawing_part = "xl/" + m.group(1)
        drawing_rels_part = ("xl/drawings/_rels/"
                             + os.path.basename(drawing_part) + ".rels")

    drawing = parts[drawing_part].decode()
    drels = parts[drawing_rels_part]

    # Where the existing captions stop, so new ones start below rather than over them.
    rows_used = [int(r) for r in re.findall(r'<row r="(\d+)"', parts[charts_part].decode())]
    next_row = (max(rows_used) if rows_used else 1) + ROW_STEP

    template_monthly = tparts[f"xl/charts/chart{MONTHLY_TEMPLATE}.xml"].decode()

    jobs = []
    for tnum, caption in VARIANTS:
        jobs.append(("copy", tnum, caption, None, None))
    for country in cfg.COUNTRY_ORDER:
        for tech in cvb.CHARTED.get(country, []):
            name = cfg.COUNTRIES[country]["name"]
            caption = (f"{name} — {tech}: capture as % of base price, by month "
                       f"(one line per year)")
            jobs.append(("monthly", None, caption, country, tech))

    for tnum, stem, src, caption in UK_VARIANTS:
        jobs.append(("variant", tnum, caption, stem, src))

    hcols = _hydro_columns()
    hydro_template = None
    if hcols and os.path.exists(HYDRO_TEMPLATE):
        with zipfile.ZipFile(HYDRO_TEMPLATE) as hz:
            hydro_template = hz.read(HYDRO_TEMPLATE_CHART).decode()
        for label, caption in _hydro_labels():
            if f"{label}_min" not in hcols:
                # A zone with no data at all publishes no columns, so there is nothing
                # to chart. Saying so beats an exhibit with three empty series.
                print(f"  hydro: no data for {label} — chart skipped", flush=True)
                continue
            jobs.append(("hydro", None, caption, label, None))
    elif not hcols:
        print("  hydro: hydro_window.csv not found — hydro charts skipped", flush=True)

    cap_rows = []
    anchors = []
    shape_id = 9000
    for k, (kind, tnum, caption, country, tech) in enumerate(jobs):
        cnum = _next_chart_num(parts)
        cpart = f"xl/charts/chart{cnum}.xml"
        if kind == "copy":
            xml = tparts[f"xl/charts/chart{tnum}.xml"].decode()
            mapping = {}
            for i, ot in enumerate(re.findall(r"<c:tx><c:strRef><c:f>([^<]+)</c:f>", xml)):
                if i < len(SERIES_NAME_CELLS):
                    mapping[ot] = SERIES_NAME_CELLS[i]
            xml = _strip_user_shapes(_repoint(xml, mapping))
        elif kind == "variant":
            xml = build_country_variant(
                tparts[f"xl/charts/chart{tnum}.xml"].decode(), country, tech, "GB")
            if xml is None:
                print(f"  skipped (no {country}.csv): {caption}", flush=True)
                continue
        elif kind == "hydro":
            xml = build_hydro_chart(hydro_template, country, hcols)
        else:
            xml = build_monthly_chart(template_monthly, country, tech, index)
        parts[cpart] = xml.encode()
        order.append(cpart)
        add_content_type(parts, cpart,
                         "application/vnd.openxmlformats-officedocument.drawingml.chart+xml")

        rid = f"rId{next_rel_id(drels)}"
        drels = drels.replace(
            b"</Relationships>",
            (f'<Relationship Id="{rid}" Type="{R}/chart" '
             f'Target="../charts/chart{cnum}.xml"/></Relationships>').encode())
        parts[drawing_rels_part] = drels

        col = CAPTION_COLS[k % 2]
        row = next_row + (k // 2) * ROW_STEP
        anchors.append(_anchor(col, row, rid, shape_id + k, f"Chart {cnum}"))
        cap_rows.append(_caption_row(row, get_column_letter(col + 1), caption))

    drawing = drawing.replace("</xdr:wsDr>", "".join(anchors) + "</xdr:wsDr>")
    parts[drawing_part] = drawing.encode()

    sheet = parts[charts_part].decode()
    sheet = sheet.replace("</sheetData>", "".join(cap_rows) + "</sheetData>")
    parts[charts_part] = sheet.encode()
    return len(jobs)


def main():
    if not os.path.exists(WB):
        raise SystemExit(f"{WB} not found — run the workbook build first")
    parts, order = read_parts(WB)

    if sheet_part_for(parts, SHEET_NAME):
        print(f"{SHEET_NAME} already present — nothing to do", flush=True)
        return

    # The two query targets come first, so their tabs exist before add_power_queries
    # runs and can be wired like every other load target. CaptureVsBase needs no query:
    # every cell on it is a formula over tabs that do.
    for tab in (EXTRA_CAPTURE_TAB, HYDRO_TAB):
        if not sheet_part_for(parts, tab):
            append_sheet(parts, order, tab, empty_sheet_xml())

    part = append_sheet(parts, order, SHEET_NAME, build_capture_vs_base_xml())
    set_full_calc_on_load(parts)
    hdr, index = cvb.layout()

    tparts, _ = read_parts(TEMPLATE)
    charts_part = sheet_part_for(parts, "Charts")
    added = add_charts(parts, order, tparts, charts_part, index)

    # The cross-market charts gain a SERIES rather than a chart, one per market beyond
    # the original five. Failing loudly here rather than skipping: a cross-market chart
    # quietly missing a country is the fault this whole block exists to correct.
    for code in cfg.COUNTRY_ORDER:
        if code in cfg.LEGACY_CSV_COUNTRIES:
            continue
        label = cfg.COUNTRIES[code]["name"]
        for part_name, sheet, header, r0, r1 in CROSS_MARKET_SERIES:
            header = header.replace("GB", code)
            problem = append_country_series(parts, part_name, sheet, header,
                                            r0, r1, label)
            if problem:
                raise SystemExit(f"cross-market series ({label}): {problem}")
        print(f"added a {label} series to {len(CROSS_MARKET_SERIES)} cross-market chart(s)",
              flush=True)

    write_parts(WB, parts, order)
    print(f"added {SHEET_NAME} ({len(hdr)} columns, {len(cvb.MONTHS)} months) -> {part}",
          flush=True)
    print(f"added {added} charts to the Charts tab", flush=True)


if __name__ == "__main__":
    main()
