"""
fetch_uk.py — pull the GB series from Elexon (+ ECB FX, + DUKES capacity) and store
them in EXACTLY the raw-parquet shapes fetch.py produces for the ENTSO-E countries.

WHY THIS FILE EXISTS AT ALL. Great Britain stopped publishing to the ENTSO-E
Transparency Platform on 15 June 2021, when the post-Brexit Trade and Cooperation
Agreement removed the obligation. Probed 2026-08-25, GB (10YGB----------A) returns:

    day-ahead price   2019-2020 only, and in GBP
    generation type   2019 to mid-2021
    installed capacity 2019 to mid-2021
    actual total load 2019 to mid-2021
    cross-border flow STILL CURRENT (the EU side of each border publishes it)
    water reservoirs  never

A TRAP WORTH KNOWING ABOUT. The "UK" domain 10Y1001A1001A92E still returns
generation and load today, so it looks like a working substitute. It is Northern
Ireland alone: 857 MW mean generation in June 2025 against 24,716 MW in June 2019.
Using it would publish NI data under a UK label. Do not.

WHAT REPLACES IT. Elexon's Insights platform publishes the very same EMFIP data
items GB stopped sending (its document ids are literally NGET-EMFIP-*), with the
same psrType vocabulary, so config.TECH_MAP needs no GB special case. No API key.

  price     MID, dataProvider APXMIDP     half-hourly, GBP/MWh, 2018 ->
  gen/type  AGPT (B1620)                  half-hourly, 2019 ->
  load      INDO                          half-hourly, 2019 ->
  capacity  NOT Elexon                    see the DUKES note below

THE PRICE IS NOT A DAY-AHEAD AUCTION, AND THAT IS A DELIBERATE, RECORDED CHOICE
(Fred, 2026-08-25). The other five markets use the ENTSO-E day-ahead auction. GB's
equivalent is N2EX, and N2EX is not obtainable free: Elexon returns N2EXMIDP rows
with price 0 and volume 0 in every year tested (2019, 2022, 2026); Nord Pool's data
portal answers 401 without a subscription; Fraunhofer's energy-charts has no GB
bidding zone; neither NESO nor Ofgem publishes a wholesale price series; EPEX's GB
auction exists only as rendered web pages. MID from APX is a within-day market index
taken near gate closure. It is the only free GB wholesale price with history, and
every surface that shows it says so. PRICE_PROVIDER below is the one-line switch if
a day-ahead feed is ever licensed.

CAPACITY COMES FROM DUKES, NOT ELEXON. Elexon's IGCA (B1410) exists but holds two
publications in its whole history, covering 2024 and 2025 — no use to a chart that
plots a run of years. DUKES 5.12.A gives Great Britain installed capacity by
generator type, 2011 to 2025, transmission and distribution. Note 5.7 is the WRONG
table despite its promising name: it de-rates wind, solar and small-scale hydro, so
it is not comparable with ENTSO-E installed capacity.

FLOWS STILL COME FROM ENTSO-E. GB's borders are published by the counterparty TSO,
so query_physical_crossborder_allborders("GB", ...) works today and is used as-is.

Usage:
  python fetch_uk.py                 # all years
  python fetch_uk.py --years 2026    # one year
  python fetch_uk.py --force         # re-fetch even if cached
"""
from __future__ import annotations
import argparse, io, json, os, re, sys, time, warnings
warnings.filterwarnings("ignore")
import pandas as pd
import requests

import config as cfg

UK = "GB"                       # the country key used throughout the pipeline
ELEXON = "https://data.elexon.co.uk/bmrs/api/v1"
ECB = "https://data-api.ecb.europa.eu/service/data/EXR/D.GBP.EUR.SP00.A"
DUKES_512 = ("https://assets.publishing.service.gov.uk/media/"
             "6a6a3668862aaf18d9c629ed/DUKES_5.12.xlsx")

# The GB price benchmark. "APXMIDP" is the only populated one (see the header note).
PRICE_PROVIDER = "APXMIDP"

# How the GB price is described everywhere it surfaces. build_status writes this to the
# Status sheet and the chart captions read it, so the basis difference cannot be lost.
PRICE_BASIS_NOTE = ("GB price is the Elexon market index (APX), a within-day index near "
                    "gate closure, not a day-ahead auction like the other markets.")

HTTP_TIMEOUT = 300
RETRIES = 4
RETRY_WAIT = 15
LOG = []


def log(m):
    print(m, flush=True)
    LOG.append(m)


def _get(url, params, timeout=HTTP_TIMEOUT, accept="application/json"):
    """GET with the same transient-status retry policy fetch.py applies to ENTSO-E.

    `accept` is a parameter because the ECB honours the Accept header over its own
    ?format= query parameter: asking for csvdata while sending accept: application/json
    returns SDMX-JSON, and the CSV parse then fails on a missing OBS_VALUE column.
    """
    wait = RETRY_WAIT
    last = None
    for attempt in range(RETRIES):
        try:
            r = requests.get(url, params=params, timeout=timeout,
                             headers={"accept": accept})
            if r.status_code == 200:
                return r
            if r.status_code not in (500, 502, 503, 504, 408, 429):
                return r                      # a real answer, even if it is "no"
            last = f"HTTP {r.status_code}"
        except Exception as ex:               # noqa: BLE001 - retry any transport fault
            last = f"{type(ex).__name__}: {ex}"
        if attempt < RETRIES - 1:
            log(f"   transient ({last}) — retry in {wait}s")
            time.sleep(wait)
            wait *= 2
    raise RuntimeError(f"{url} failed after {RETRIES} attempts: {last}")


def raw_path(series, year):
    return os.path.join(cfg.RAW_DIR, f"{UK}_{series}_{year}.parquet")


# THE TRAILING WINDOW, added 2026-08-26. Until then this script pulled TWO WHOLE YEARS on
# every run, because the workflow called it `--years 2025,2026 --force` and it had no other
# mode. That was invisible while ENTSO-E's own slowness dwarfed it; once the other five
# markets dropped to between three and nine minutes, GB was the longest fetch in the stage.
#
# Same rule as fetch.py: a window is only safe to merge into something that already exists,
# so a series with nothing stored widens back to the whole year, per series rather than per
# run. That per-series granularity is not fussiness. On 2026-07-31 a country-level check
# let a series whose earlier fetch had failed take the incremental path and write a 30-day
# "year", and the same shape of fault cost France and Italy six months.
SINCE_DAYS = None
_ACTIVE = None          # (start, end) for the series being fetched, or None for a full year


# THE GAPS RECORD, added 2026-08-26, and the reason is worth stating plainly: until today
# this script COULD NOT FAIL. `main` returned None on every path and `fetch_year` caught
# every exception, so a total Elexon outage exited 0 with a cheerful log. Three things
# followed from that, all of them silent.
#
#   * The workflow's own safety net was unreachable. `if ! python fetch_uk.py --since-days
#     45; then ... --force` can only fire on a non-zero exit, and there was none.
#   * Nothing wrote `fetch-gaps.json`, which is the ONLY route by which a fetch problem
#     reaches a human: the publish job builds health.json from it, the status page reads
#     that, and build_status.py turns it into the Excel banner naming the affected tabs.
#     Great Britain could not appear in any of them.
#   * There was no bound on how old a stored fallback could be. The other five markets
#     have had one since 2026-08-23.
#
# Downstream would not have caught it either. check_coverage looks for data that SHRANK,
# and a frozen GB column loses nothing while the other five keep the file's last row
# advancing. So this is the whole of GB's failure reporting, and it is deliberately the
# same shape as fetch.py's: same file name, same record fields, same three outcomes, so
# the publish job, the repair workflow and the status sheet need no GB special case.
REQUIRED = ("price", "load", "generation")
FALLBACK_DAYS = 3               # matches fetch.py; see the note on FALLBACK_DAYS there
GAPS_FILE = "fetch-gaps.json"

# path -> (label, "ok" | "fail" | "skip"). "skip" is a completed year already stored, whose
# age says nothing about health, exactly as in fetch.py.
OUTCOMES = {}


def _mark(path, label, state):
    """Record an outcome. A later success overwrites an earlier failure; not the reverse."""
    if state == "ok" or path not in OUTCOMES:
        OUTCOMES[path] = (label, state)


def _required(label):
    return any(label.startswith(r) for r in REQUIRED)


def _stored_coverage_end(path):
    """How far the STORED copy runs, from the data's own last timestamp.

    Not the file mtime: a cache restore rewrites that, so a file whose data stops in June
    can look like it was written this morning.
    """
    try:
        df = pd.read_parquet(path)
        return pd.Timestamp(df.index.max()).tz_convert("UTC") if len(df.index) else None
    except Exception:                                     # noqa: BLE001 - unreadable
        return None


def classify_gaps():
    """Required series that failed, split into what the run survives and what it cannot."""
    hard, stale = [], []
    now = pd.Timestamp.now(tz="UTC")
    for path, (label, state) in OUTCOMES.items():
        if state != "fail" or not _required(label):
            continue
        base = os.path.basename(path)
        if not (os.path.exists(path) and os.path.getsize(path) > 0):
            hard.append({"series": label, "file": base, "why": "nothing stored"})
            continue
        m = re.search(r"_(\d{4})\.parquet$", path)
        year = int(m.group(1)) if m else None
        if year is not None and year < cfg.CURRENT_YEAR:
            continue                                      # a completed year is complete
        end = _stored_coverage_end(path)
        if end is None:
            hard.append({"series": label, "file": base, "why": "stored file unreadable"})
            continue
        age = (now - end).days
        if age > FALLBACK_DAYS:
            hard.append({"series": label, "file": base,
                         "why": f"stored data ends {end.date()}, {age} days old, past the "
                                f"{FALLBACK_DAYS}-day bound"})
        else:
            stale.append({"series": label, "file": base,
                          "covers_to": end.isoformat(timespec="minutes"), "days_old": age})
    return hard, stale


def write_gaps(hard, stale, years):
    """Same record fields fetch.py writes, so every consumer reads GB with no special case."""
    if not (hard or stale):
        return
    rec = {"at": pd.Timestamp.now(tz="UTC").isoformat(timespec="seconds"),
           "countries": [UK], "years": sorted(years),
           "fatal": hard, "stale": stale,
           "series": sorted({g["series"] for g in hard + stale})}
    try:
        os.makedirs(cfg.META_DIR, exist_ok=True)
        with io.open(os.path.join(cfg.META_DIR, GAPS_FILE), "w",
                     encoding="utf-8", newline="\n") as fh:
            json.dump(rec, fh, indent=1)
        log(f"  wrote {GAPS_FILE}: {len(hard)} fatal, {len(stale)} running on stored data")
    except Exception as ex:                               # noqa: BLE001
        log(f"  could not write {GAPS_FILE}: {ex}")


def _need(path, force):
    if SINCE_DAYS is not None:
        return True     # a windowed run always re-fetches; the merge keeps the history
    return force or not (os.path.exists(path) and os.path.getsize(path) > 0)


def _begin(path, year):
    """Decide this series' window and remember it for _stream and _save. -> (start, end)."""
    global _ACTIVE
    _ACTIVE = None
    ys = pd.Timestamp(f"{year}-01-01", tz="UTC")
    ye = pd.Timestamp(f"{year + 1}-01-01", tz="UTC")
    if SINCE_DAYS is None:
        return ys, ye
    if not (os.path.exists(path) and os.path.getsize(path) > 0):
        log(f"   (nothing stored for {os.path.basename(path)} — full year, not the window)")
        return ys, ye
    cut = pd.Timestamp.now(tz="UTC").floor("h") - pd.Timedelta(days=int(SINCE_DAYS))
    start = max(ys, cut)
    if start <= ys:
        return ys, ye                       # the window already covers the whole year
    _ACTIVE = (start, ye)
    return start, ye


def _merge_into(path, fresh):
    """Fresh rows win on overlap; anything outside the window is left alone.

    Elexon restates settlement periods, so a re-fetched half-hour must REPLACE the stored
    one rather than be discarded as a duplicate. That is why the window exists at all.
    """
    if fresh is None or len(fresh) == 0:
        return None
    if not (os.path.exists(path) and os.path.getsize(path) > 0):
        return fresh
    old = pd.read_parquet(path)
    old.index = pd.DatetimeIndex(old.index).tz_convert("UTC")
    fresh = fresh.copy()
    fresh.index = pd.DatetimeIndex(fresh.index).tz_convert("UTC")
    both = pd.concat([old, fresh])
    both = both[~both.index.duplicated(keep="last")].sort_index()
    return both


def _save(df, path, label, year=None, expect_halfhourly=False):
    """Write in fetch.py's stored shape: UTC index named ts_utc, string columns.

    `expect_halfhourly` turns on a ROW-COUNT FLOOR for a completed year, and it is here
    because of a fault it would have caught on the day it was written: the load pull
    looped over months against an endpoint that caps at seven days, kept only the one
    month short enough to succeed, and wrote a 1,392-row year that every later step
    accepted as a year. Short is not the same as absent, and nothing downstream can tell
    the difference. A fetch that came back thin has to say so, in its own step.
    """
    if df is None or len(df) == 0:
        log(f"   {label}: NO DATA — not written")
        _mark(path, label, "fail")
        return False
    df = df.copy()
    df.index = pd.DatetimeIndex(df.index).tz_convert("UTC")
    # A WINDOWED FETCH MERGES; a full one replaces. Writing a 45-day frame straight over a
    # stored year is exactly the "wrote a 1,392-row year" fault this function's own docstring
    # warns about, one step earlier, so the merge happens BEFORE the row-count floor below
    # and the floor still sees a whole year.
    if _ACTIVE is not None:
        merged = _merge_into(path, df)
        if merged is not None:
            df = merged
    df.index.name = "ts_utc"
    df.columns = [str(c) for c in df.columns]
    df = df[~df.index.duplicated(keep="first")].sort_index()
    df.to_parquet(path)
    note = ""
    if expect_halfhourly and year is not None and year < cfg.CURRENT_YEAR:
        expected = 366 * 48 if year % 4 == 0 else 365 * 48
        if len(df) < 0.9 * expected:
            note = f"  ** WARN: {len(df)} rows against ~{expected} expected for a full year"
    log(f"   {label}: {len(df)} rows -> {os.path.basename(path)}{note}")
    _mark(path, label, "ok")
    return True


# ---------------------------------------------------------------------------
# FX — daily GBP per EUR, from the ECB reference rates
# ---------------------------------------------------------------------------
_FX_CACHE = None


def fx_gbp_per_eur():
    """Daily GBP-per-EUR, forward-filled onto a calendar-day index.

    The ECB publishes on TARGET business days only, so weekends and holidays are
    absent. Forward-filling carries the last published rate across them, which is
    what a market participant would actually have transacted at. Reindexing without
    the fill would silently drop every weekend hour from the converted price series.
    """
    global _FX_CACHE
    if _FX_CACHE is not None:
        return _FX_CACHE
    r = _get(ECB, {"format": "csvdata", "startPeriod": f"{cfg.START_YEAR - 1}-01-01"},
             accept="text/csv")
    if r.status_code != 200:
        raise RuntimeError(f"ECB FX fetch failed: HTTP {r.status_code}")
    df = pd.read_csv(io.StringIO(r.text))
    s = pd.Series(df["OBS_VALUE"].values,
                  index=pd.to_datetime(df["TIME_PERIOD"]).values).sort_index()
    full = pd.date_range(s.index.min(), pd.Timestamp.today().normalize(), freq="D")
    _FX_CACHE = s.reindex(full).ffill()
    log(f"   FX: {len(s)} ECB observations, {s.index.min().date()}..{s.index.max().date()}")
    return _FX_CACHE


def to_eur(gbp_series):
    """GBP/MWh -> EUR/MWh at that day's ECB rate (rate is GBP per EUR, so divide)."""
    fx = fx_gbp_per_eur()
    days = pd.DatetimeIndex(gbp_series.index).tz_convert("UTC").normalize().tz_localize(None)
    rate = pd.Series(fx.reindex(days).values, index=gbp_series.index)
    if rate.isna().any():
        # Only the current day can legitimately outrun the ECB publication.
        rate = rate.ffill()
    return gbp_series / rate


# ---------------------------------------------------------------------------
# Elexon pulls
# ---------------------------------------------------------------------------
def _stream(dataset, year, time_from_key, time_to_key, extra=None):
    """One whole year in one request via the /stream variant.

    The plain /datasets endpoints cap a MID query at seven days, which would make a
    backfill several hundred calls. The /stream variants take a full year, verified
    2026-08-25 (MID/stream returned 17,521 rows for 2025, a complete half-hourly year).
    """
    if _ACTIVE is not None:
        a, b = _ACTIVE
        params = {time_from_key: a.strftime("%Y-%m-%dT%H:%MZ"),
                  time_to_key: b.strftime("%Y-%m-%dT%H:%MZ")}
    else:
        params = {time_from_key: f"{year}-01-01T00:00Z",
                  time_to_key: f"{year + 1}-01-01T00:00Z"}
    params.update(extra or {})
    r = _get(f"{ELEXON}/datasets/{dataset}/stream", params)
    if r.status_code != 200:
        log(f"   {dataset} {year}: HTTP {r.status_code} {r.text[:120]}")
        return pd.DataFrame()
    d = r.json()
    if isinstance(d, dict):
        d = d.get("data", [])
    return pd.DataFrame(d)


def fetch_price(year, force):
    """GB wholesale price, half-hourly, converted GBP -> EUR. Stored as price_GB."""
    path = raw_path(f"price_{UK}", year)
    if not _need(path, force):
        log(f"   price: cached")
        _mark(path, "price", "skip")
        return
    _begin(path, year)
    df = _stream("MID", year, "from", "to", {"dataProviders": PRICE_PROVIDER})
    if df.empty:
        _save(None, path, "price")
        return
    df["ts"] = pd.to_datetime(df["startTime"], utc=True)
    s = df.set_index("ts")["price"].sort_index()
    # A zero here is a non-report, not a zero clearing price: the whole reason N2EX is
    # unusable is that it reports 0.0 for every period. Dropping them keeps a provider
    # outage out of the price statistics rather than dragging the annual mean down.
    s = s[s != 0.0]
    _save(to_eur(s).to_frame(name="value"), path, "price", year, True)


def fetch_load(year, force):
    """GB national demand outturn, half-hourly MW.

    INDO, not ATL. ATL is the exact ENTSO-E item (B0610) but Elexon only holds it for
    recent months, whereas INDO runs from 2019 to today (probed 2026-08-25).

    ⚠ THIS IS TRANSMISSION DEMAND, AND IT IS NOT THE SAME BASIS AS ENTSO-E'S LOAD.
    INDO is metered at the transmission boundary, so it is already NET of Britain's
    distribution-connected generation — most of which is solar. ENTSO-E's figures for the
    other five markets are not netted that way. GENERATION is unaffected: Elexon's AGPT
    does include embedded plant, verified 2026-08-25 by GB solar peaking at 13.5 GW
    against 21.5 GW installed, a load factor only reachable if embedded solar is counted.

    So GB load must NEVER be dropped into a net-load calculation of the form
    (demand - wind - solar) without correcting the basis: embedded solar would be
    subtracted twice, deepening the midday trough by roughly its own size. No exhibit
    does this today — the net-load duck is Germany only, and GB load currently feeds no
    chart at all — which is exactly why this warning is here rather than in a fix.
    """
    path = raw_path("load", year)
    if not _need(path, force):
        log(f"   load: cached")
        _mark(path, "load", "skip")
        return
    _begin(path, year)
    # The plain endpoint caps at seven days and returns HTTP 400 beyond it, so a loop
    # over months silently kept only the one month that happened to be short enough and
    # wrote a 1,392-row "year". The /stream variant takes the whole year in one call
    # (17,568 half-hourly rows for 2025, verified 2026-08-25).
    if _ACTIVE is not None:
        _a, _b = _ACTIVE
        _dates = {"settlementDateFrom": _a.strftime("%Y-%m-%d"),
                  "settlementDateTo": _b.strftime("%Y-%m-%d")}
    else:
        _dates = {"settlementDateFrom": f"{year}-01-01",
                  "settlementDateTo": f"{year + 1}-01-01"}
    r = _get(f"{ELEXON}/demand/outturn/stream", _dates)
    if r.status_code != 200:
        log(f"   load: HTTP {r.status_code} {r.text[:120]}")
        _save(None, path, "load")
        return
    d = r.json()
    df = pd.DataFrame(d if isinstance(d, list) else d.get("data", []))
    if df.empty:
        _save(None, path, "load")
        return
    df["ts"] = pd.to_datetime(df["startTime"], utc=True)
    col = "initialTransmissionSystemDemandOutturn" if \
        "initialTransmissionSystemDemandOutturn" in df.columns else "initialDemandOutturn"
    s = df.set_index("ts")[col].sort_index()
    _save(s.to_frame(name="Actual Load"), path, "load", year, True)


def fetch_generation(year, force):
    """GB generation per production type, half-hourly, in ENTSO-E's own psrType names.

    Stored with fetch.py's "psr|business_type" column convention so build_hourly needs
    no GB branch. Elexon's businessType is a reporting label ("Production", "Wind
    generation", "Solar generation") and every one of them is production, so they all
    map to "Actual Aggregated".

    ONE HONEST GAP: AGPT carries no Actual Consumption rows, so GB has no
    pumped-storage CONSUMPTION series. The intraday mix chart shows that as a negative
    band for the other markets; for GB the band is simply absent rather than zero.
    """
    path = raw_path("generation", year)
    if not _need(path, force):
        log(f"   generation: cached")
        _mark(path, "generation", "skip")
        return
    _begin(path, year)
    df = _stream("AGPT", year, "publishDateTimeFrom", "publishDateTimeTo")
    if df.empty:
        _save(None, path, "generation")
        return
    df["ts"] = pd.to_datetime(df["startTime"], utc=True)
    # A restated settlement period appears again with a higher revision number.
    if "documentRevisionNumber" in df.columns:
        df = (df.sort_values("documentRevisionNumber")
                .drop_duplicates(["ts", "psrType"], keep="last"))
    wide = df.pivot_table(index="ts", columns="psrType", values="quantity", aggfunc="last")
    wide.columns = [f"{c}|Actual Aggregated" for c in wide.columns]
    _save(wide.sort_index(), path, "generation", year, True)


def fetch_flows(year, force):
    """Cross-border physical flows, still on ENTSO-E because the EU side publishes them."""
    ip, ep = raw_path("flow_import", year), raw_path("flow_export", year)
    if not _need(ip, force) and not _need(ep, force):
        log(f"   flows: cached")
        _mark(ip, "flow_import", "skip"); _mark(ep, "flow_export", "skip")
        return
    from entsoe import EntsoePandasClient
    import crossborder
    if not cfg.API_KEY:
        log("   flows: no ENTSO-E key — skipped")
        return

    def _client():
        # One per worker: EntsoePandasClient wraps a requests.Session, which is not
        # guaranteed thread-safe. Same settings as the serial client it replaces.
        return EntsoePandasClient(api_key=cfg.API_KEY, retry_count=4, retry_delay=8)

    now = pd.Timestamp.now(tz="UTC").floor("h")
    for label, path, is_export in (("flow_import", ip, False), ("flow_export", ep, True)):
        if not _need(path, force):
            continue
        # Per series, so a flow file that is missing widens to the year while the other
        # takes the window.
        s, e = _begin(path, year)
        e = min(e, now)
        if s >= e:
            continue
        try:
            # CONCURRENTLY, matching what fetch.py does for the other five markets. GB was
            # left on the library's own helper, which makes one request per neighbour in
            # sequence; crossborder_test drives both over the same canned responses and
            # asserts the frames are equal, column order included.
            df = crossborder.all_borders(_client, UK, start=s, end=e, export=is_export)
            _save(df, path, label)
        except Exception as ex:               # noqa: BLE001 - a border gap is not fatal
            log(f"   {label}: {type(ex).__name__} — not written")
            _mark(path, label, "fail")


# ---------------------------------------------------------------------------
# Capacity — DUKES 5.12.A
# ---------------------------------------------------------------------------
# DUKES generator type -> ENTSO-E psrType, so build_hourly's TECH_MAP does the rest.
# Nuclear arrives as three reactor families and is summed. "OCGT and conventional
# thermal gas" is gas-fired plant that is not CCGT, so it joins Fossil Gas rather than
# becoming Other, which would have put roughly 3 GW in an unlabelled bucket.
DUKES_TO_PSR = {
    "Coal": "Fossil Hard coal",
    "Oil": "Fossil Oil",
    "OCGT and conventional thermal gas": "Fossil Gas",
    "CCGT gas": "Fossil Gas",
    "Nuclear - Magnox": "Nuclear",
    "Nuclear - PWR": "Nuclear",
    "Nuclear - AGR": "Nuclear",
    "Hydro (natural flow)": "Hydro Run-of-river and poundage",
    "Wind (onshore)": "Wind Onshore",
    "Wind (offshore)": "Wind Offshore",
    "Solar": "Solar",
    "Wave and tidal": "Marine",
    "Bioenergy": "Biomass",
    "Other fuels": "Other",
    "Pumped storage": "Hydro Pumped Storage",
}


def _dukes_label(v):
    """'Coal [note 4]' -> 'Coal'."""
    return str(v).split("[")[0].strip()


def fetch_capacity(years, force):
    """Installed capacity by technology from DUKES 5.12.A (Great Britain).

    Written one file per year in ENTSO-E's capacity shape — a single-row frame whose
    columns are psrType names — so build_hourly.build_capacity consumes it unchanged.

    Table 5.12.B on the same sheet is Northern Ireland and is deliberately NOT added:
    the GB price, generation and load series are all Great Britain, so folding NI into
    capacity alone would make the capacity chart describe a different system from every
    other UK exhibit.
    """
    targets = [y for y in years if _need(raw_path("capacity", y), force)]
    if not targets:
        log("   capacity: cached")
        return
    import openpyxl
    local = os.path.join(cfg.RAW_DIR, "DUKES_5.12.xlsx")
    # RAISE, DO NOT RETURN. This URL embeds a content hash that DESNZ rotates every time
    # it re-issues DUKES, so it WILL stop resolving, probably at the next annual release.
    # Logging and returning meant GB installed capacity would quietly freeze at its last
    # stored year: check_coverage sees no shrink (the old parquet is still there), the
    # chart keeps drawing, and nothing anywhere says the series stopped advancing. A
    # source that has moved is news on the day it moves.
    r = _get(DUKES_512, {}, accept="*/*")
    if r.status_code != 200:
        raise RuntimeError(
            f"DUKES capacity download failed: HTTP {r.status_code} for {DUKES_512}. "
            f"The gov.uk media id rotates on re-issue — find the current DUKES 5.12 "
            f"link from the electricity chapter page and update DUKES_512.")
    with open(local, "wb") as f:
        f.write(r.content)
    ws = openpyxl.load_workbook(local, data_only=True)["5.12"]

    # Locate 5.12.A's header row rather than hardcoding it, so a re-issue that adds a
    # note line above the table does not silently shift every year by one column.
    hdr = None
    for row in range(1, 40):
        if str(ws.cell(row, 1).value).strip() == "Network type":
            hdr = row
            break
    if hdr is None:
        raise RuntimeError(
            "DUKES 5.12: could not find the 'Network type' header row. The workbook's "
            "layout has changed; re-read it before trusting any capacity figure.")
    year_col = {}
    for c in range(3, ws.max_column + 1):
        v = ws.cell(hdr, c).value
        if v is None:
            continue
        try:
            year_col[int(str(v).strip())] = c
        except ValueError:
            continue

    # 5.12.A only: stop at the Northern Ireland table.
    end = hdr + 1
    while end <= ws.max_row and str(ws.cell(end, 1).value).strip() in ("Transmission",
                                                                      "Distribution"):
        end += 1

    written = 0
    for y in targets:
        c = year_col.get(y)
        if c is None:
            log(f"   capacity {y}: not in DUKES yet — not written")
            continue
        agg = {}
        for row in range(hdr + 1, end):
            label = _dukes_label(ws.cell(row, 2).value)
            if label.startswith("Total"):
                continue                      # a subtotal row, would double-count
            psr = DUKES_TO_PSR.get(label)
            if psr is None:
                continue
            val = ws.cell(row, c).value
            if isinstance(val, (int, float)):
                agg[psr] = agg.get(psr, 0.0) + float(val)
        if not agg:
            log(f"   capacity {y}: no rows matched — not written")
            continue
        frame = pd.DataFrame([agg], index=pd.DatetimeIndex([pd.Timestamp(f"{y}-01-01",
                                                                        tz="UTC")]))
        _save(frame, raw_path("capacity", y), f"capacity {y}")
        written += 1
    log(f"   capacity: {written} year(s) from DUKES 5.12.A")


# ---------------------------------------------------------------------------
def fetch_year(year, force):
    """Each series is attempted independently, and a raise is RECORDED rather than lost.

    Before 2026-08-26 an exception here escaped to `main`, which logged it and carried on
    to exit 0. Catching per series keeps that resilience — one dead endpoint should not
    cost the other five — while leaving a mark that classify_gaps can act on.
    """
    log(f"== {UK} (Elexon + ECB + DUKES) {year} ==")
    for fn, label, path in ((fetch_price, "price", raw_path(f"price_{UK}", year)),
                            (fetch_load, "load", raw_path("load", year)),
                            (fetch_generation, "generation", raw_path("generation", year))):
        try:
            fn(year, force)
        except Exception as ex:                # noqa: BLE001 - recorded, not swallowed
            log(f"   {label}: {type(ex).__name__}: {ex} — not written")
            _mark(path, label, "fail")
    try:
        fetch_flows(year, force)
    except Exception as ex:                    # noqa: BLE001
        log(f"   flows: {type(ex).__name__}: {ex} — not written")
        for lbl in ("flow_import", "flow_export"):
            _mark(raw_path(lbl, year), lbl, "fail")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--years", default=None, help="comma list, e.g. 2026")
    ap.add_argument("--force", action="store_true")
    ap.add_argument("--since-days", type=int, default=None,
                    help="fetch only the trailing N days and merge into stored data "
                         "(falls back to a full year, per series, if nothing is stored)")
    a = ap.parse_args()
    years = [int(y) for y in a.years.split(",")] if a.years else cfg.YEARS
    global SINCE_DAYS
    SINCE_DAYS = a.since_days
    if SINCE_DAYS:
        log(f"trailing window: last {SINCE_DAYS} days, merging into stored data")

    t0 = time.time()
    for y in years:
        try:
            fetch_year(y, a.force)
        except Exception as ex:               # noqa: BLE001 - report and keep going
            log(f"UNCAUGHT {UK} {y}: {type(ex).__name__}: {ex}")
            for lbl, key in (("price", f"price_{UK}"), ("load", "load"),
                             ("generation", "generation")):
                _mark(raw_path(key, y), lbl, "fail")
    try:
        fetch_capacity(years, a.force)
    except Exception as ex:                   # noqa: BLE001 - capacity is not required
        log(f"UNCAUGHT {UK} capacity: {type(ex).__name__}: {ex}")
    log(f"DONE in {(time.time() - t0) / 60:.1f} min")
    with open(os.path.join(cfg.META_DIR, "fetch_uk.log"), "w", encoding="utf-8",
              newline="\n") as f:
        f.write("\n".join(LOG) + "\n")

    # A FETCH THAT DID NOT FETCH FAILS HERE, in its own step, rather than surfacing 25
    # minutes downstream as a chart-geometry complaint that says nothing about Elexon.
    hard, stale = classify_gaps()
    write_gaps(hard, stale, years)
    for g in stale:
        log(f"  STALE {g['series']}: fetch failed, continuing on stored data to "
            f"{g['covers_to']} ({g['days_old']}d old, bound is {FALLBACK_DAYS}d)")
    if hard:
        for g in hard:
            log(f"  MISSING {g['series']}: {g['why']}")
        print("\n".join(LOG[-40:]))
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
